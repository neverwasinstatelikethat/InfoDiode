"""Тесты для параметрического маппера — rigorous validation of parameter table loading.

Tests verify loading from Excel (.xlsx) and CSV formats, adaptive column detection,
sheet handling, and graceful error handling for malformed data.
"""

import pytest
from pathlib import Path
from app.core.parameter_mapper import (
    load_parameter_table,
    _detect_header_row,
    _detect_column_roles,
    _find_name_column_by_content,
)


class TestParameterMapperExcel:
    """Тесты загрузки параметров из Excel (.xlsx)."""

    def test_load_excel_four_sheets_param_counts(self, tmp_path) -> None:
        """Тест: загрузка Excel с 4 листами → проверка точного количества параметров.
        
        Simulates the real parameter table structure with 4 sheets:
        - Sheet 1: 71 parameters
        - Sheet 2: 75 parameters  
        - Sheet 3: 79 parameters
        - Sheet 4: 36 parameters
        Total: 261 parameters
        """
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        # Create workbook with 4 sheets
        wb = Workbook()
        
        # Sheet 1: 71 temperature parameters
        ws1 = wb.active
        ws1.title = "1_ai"
        ws1.append(["№", "Наименование", "Краткое имя", "Единица", "Знаки"])
        for i in range(1, 72):
            ws1.append([i, f"Температура датчик {i}", "T", "°С", 1])
        
        # Sheet 2: 75 pressure parameters
        ws2 = wb.create_sheet("2_ai")
        ws2.append(["№", "Наименование", "Краткое имя", "Единица", "Знаки"])
        for i in range(1, 76):
            ws2.append([i, f"Давление датчик {i}", "P", "кПа", 2])
        
        # Sheet 3: 79 position parameters
        ws3 = wb.create_sheet("3_ai")
        ws3.append(["№", "Наименование", "Краткое имя", "Единица", "Знаки"])
        for i in range(1, 80):
            ws3.append([i, f"Положение задвижки {i}", "Pos", "%", 1])
        
        # Sheet 4: 36 misc parameters
        ws4 = wb.create_sheet("4_ai")
        ws4.append(["№", "Наименование", "Краткое имя", "Единица", "Знаки"])
        for i in range(1, 37):
            ws4.append([i, f"Резервный параметр {i}", "R", "", 0])
        
        filepath = tmp_path / "params_four_sheets.xlsx"
        wb.save(filepath)
        
        params = load_parameter_table(str(filepath))
        
        # Verify total count
        assert len(params) == 261
        
        # Verify each sheet's params are loaded
        sheet1_params = [p for p in params if p.get("sheet_name") == "1_ai"]
        sheet2_params = [p for p in params if p.get("sheet_name") == "2_ai"]
        sheet3_params = [p for p in params if p.get("sheet_name") == "3_ai"]
        sheet4_params = [p for p in params if p.get("sheet_name") == "4_ai"]
        
        assert len(sheet1_params) == 71
        assert len(sheet2_params) == 75
        assert len(sheet3_params) == 79
        assert len(sheet4_params) == 36

    def test_excel_sheet_name_field_populated(self, tmp_path) -> None:
        """Тест: поле sheet_name заполняется корректно для каждого параметра."""
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Temperatures"
        ws1.append(["Наименование", "Краткое имя", "Единица"])
        ws1.append(["Температура 1", "T1", "°С"])
        
        ws2 = wb.create_sheet("Pressures")
        ws2.append(["Наименование", "Краткое имя", "Единица"])
        ws2.append(["Давление 1", "P1", "кПа"])
        
        filepath = tmp_path / "multi_sheet.xlsx"
        wb.save(filepath)
        
        params = load_parameter_table(str(filepath))
        
        assert len(params) == 2
        
        # Find params by name and verify sheet_name
        temp_param = next(p for p in params if "Температура" in p["name"])
        press_param = next(p for p in params if "Давление" in p["name"])
        
        assert temp_param["sheet_name"] == "Temperatures"
        assert press_param["sheet_name"] == "Pressures"

    def test_excel_name_unit_short_name_extraction(self, tmp_path) -> None:
        """Тест: корректное извлечение name, unit, short_name из нужных колонок."""
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Parameters"
        # Headers in Russian as in real files
        ws.append(["№", "Наименование параметра", "Обозначение", "Единица измерения", "Знаки после запятой"])
        ws.append([1, "Температура газа на входе", "T_in", "°С", 1])
        ws.append([2, "Давление в баке", "P_tank", "МПа", 3])
        
        filepath = tmp_path / "params_detailed.xlsx"
        wb.save(filepath)
        
        params = load_parameter_table(str(filepath))
        
        assert len(params) == 2
        
        # Verify first param
        p1 = params[0]
        assert p1["name"] == "Температура газа на входе"
        assert p1["short_name"] == "T_in"
        assert p1["unit"] == "°С"
        assert p1["decimal_places"] == 1
        
        # Verify second param
        p2 = params[1]
        assert p2["name"] == "Давление в баке"
        assert p2["short_name"] == "P_tank"
        assert p2["unit"] == "МПа"
        assert p2["decimal_places"] == 3

    def test_adaptive_column_detection_different_headers(self, tmp_path) -> None:
        """Тест: адаптивное определение колонок с различными форматами заголовков.
        
        Verifies that column detection works with various header naming conventions.
        """
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        wb = Workbook()
        ws = wb.active
        # Different header format (synonyms)
        ws.append(["ID", "Русское имя", "Тег", "Размерность", "Точность"])
        ws.append([1, "Температура", "T", "°С", 1])
        
        filepath = tmp_path / "params_synonyms.xlsx"
        wb.save(filepath)
        
        params = load_parameter_table(str(filepath))
        
        assert len(params) == 1
        assert params[0]["name"] == "Температура"
        assert params[0]["short_name"] == "T"
        assert params[0]["unit"] == "°С"

    def test_excel_empty_rows_skipped(self, tmp_path) -> None:
        """Тест: пустые строки пропускаются при загрузке."""
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        wb = Workbook()
        ws = wb.active
        ws.append(["Наименование", "Краткое имя", "Единица"])
        ws.append(["Параметр 1", "P1", "мм"])
        ws.append([None, None, None])  # Empty row
        ws.append(["", "", ""])  # Another empty row
        ws.append(["Параметр 2", "P2", "см"])
        
        filepath = tmp_path / "params_with_empty.xlsx"
        wb.save(filepath)
        
        params = load_parameter_table(str(filepath))
        
        assert len(params) == 2
        assert params[0]["name"] == "Параметр 1"
        assert params[1]["name"] == "Параметр 2"


class TestParameterMapperCSV:
    """Тесты загрузки параметров из CSV."""

    def test_load_csv_parameter_table(self, tmp_path) -> None:
        """Тест: базовая загрузка CSV с разделителем ;"""
        csv_content = "Русское имя;Краткое имя;ID датчика;Единица;Мин;Макс\nДавление P1;P1;PT001;МПа;0;10\nТемпература T1;T1;TT001;°C;-50;150\nРасход F1;F1;FT001;м³/ч;0;100"
        csv_file = tmp_path / "params.csv"
        csv_file.write_text(csv_content, encoding="utf-8")

        params = load_parameter_table(str(csv_file))
        
        assert len(params) == 3
        
        # Verify first param - name and unit are extracted
        assert params[0]["name"] == "Давление P1"
        assert params[0]["unit"] == "МПа"
        # Note: CSV loader doesn't extract short_name (set to empty)
        assert "short_name" in params[0]
        
        # Verify second param
        assert params[1]["name"] == "Температура T1"
        assert params[1]["unit"] == "°C"

    def test_load_csv_windows1251_encoding(self, tmp_path) -> None:
        """Тест: автоматическое определение кодировки Windows-1251."""
        csv_content = "Наименование;Единица\nТемпература;°С\nДавление;кПа"
        csv_file = tmp_path / "params_1251.csv"
        csv_file.write_bytes(csv_content.encode("windows-1251"))

        params = load_parameter_table(str(csv_file))
        
        assert len(params) == 2
        assert "Температура" in params[0]["name"]
        assert "Давление" in params[1]["name"]

    def test_load_csv_utf8_encoding(self, tmp_path) -> None:
        """Тест: корректная загрузка CSV в UTF-8."""
        csv_content = "Наименование;Единица\nТемпература;°С\nДавление;кПа"
        csv_file = tmp_path / "params_utf8.csv"
        csv_file.write_text(csv_content, encoding="utf-8")

        params = load_parameter_table(str(csv_file))
        
        assert len(params) == 2
        assert params[0]["name"] == "Температура"
        assert params[0]["unit"] == "°С"


class TestParameterMapperErrors:
    """Тесты обработки ошибок и граничных случаев."""

    def test_load_nonexistent_file(self) -> None:
        """Тест: FileNotFoundError при загрузке несуществующего файла."""
        with pytest.raises(FileNotFoundError):
            load_parameter_table("nonexistent_file.xlsx")

    def test_load_empty_csv(self, tmp_path) -> None:
        """Тест: пустой CSV возвращает пустой список."""
        csv_file = tmp_path / "empty.csv"
        csv_file.write_text("", encoding="utf-8")
        params = load_parameter_table(str(csv_file))
        assert len(params) == 0

    def test_load_csv_only_headers(self, tmp_path) -> None:
        """Тест: CSV только с заголовками возвращает пустой список."""
        csv_file = tmp_path / "headers_only.csv"
        csv_file.write_text("Наименование;Единица\n", encoding="utf-8")
        params = load_parameter_table(str(csv_file))
        assert len(params) == 0

    def test_load_unsupported_format(self, tmp_path) -> None:
        """Тест: ValueError при загрузке неподдерживаемого формата."""
        txt_file = tmp_path / "params.txt"
        txt_file.write_text("some content", encoding="utf-8")
        
        with pytest.raises(ValueError, match="Неподдерживаемый формат"):
            load_parameter_table(str(txt_file))

    def test_missing_columns_graceful_handling(self, tmp_path) -> None:
        """Тест: отсутствующие колонки обрабатываются gracefully.
        
        When some columns are missing, the loader should still work
        and populate available fields.
        """
        csv_file = tmp_path / "partial.csv"
        # Only name column, no unit or short_name
        csv_file.write_text("Наименование\nТемпература 1\nДавление 1", encoding="utf-8")
        
        params = load_parameter_table(str(csv_file))
        
        assert len(params) == 2
        assert params[0]["name"] == "Температура 1"
        assert params[0]["unit"] == ""  # Empty string for missing
        assert params[0]["short_name"] == ""  # Empty string for missing


class TestColumnDetection:
    """Тесты функций определения колонок."""

    def test_detect_header_row_with_text_headers(self) -> None:
        """Тест: определение строки с текстовыми заголовками."""
        rows = [
            ["Some", "header", "text"],
            [1, "Data row", "value"],
        ]
        header_idx = _detect_header_row(rows, max_rows=5)
        assert header_idx == 0

    def test_detect_header_row_skips_empty(self) -> None:
        """Тест: пропуск пустых строк перед заголовками."""
        rows = [
            [None, None],
            ["", ""],
            ["Наименование", "Единица"],
            ["Data", "value"],
        ]
        header_idx = _detect_header_row(rows, max_rows=5)
        assert header_idx == 2

    def test_detect_column_roles_name(self) -> None:
        """Тест: определение колонки name по ключевым словам."""
        headers = ["№", "наименование параметра", "ед.изм", "знаки"]
        roles = _detect_column_roles(headers)
        
        assert roles.get("name") == 1  # "наименование параметра"
        assert roles.get("unit") == 2   # "ед.изм"

    def test_detect_column_roles_short_name(self) -> None:
        """Тест: определение колонки short_name по ключевым словам."""
        headers = ["описание", "обозначение", "размерность"]
        roles = _detect_column_roles(headers)
        
        assert roles.get("short_name") == 1  # "обозначение"

    def test_find_name_column_by_content(self) -> None:
        """Тест: fallback определение name колонки по содержимому."""
        rows = [
            ["ID", "Short", "Full parameter name here"],
            [1, "T1", "Temperature sensor number one"],
            [2, "P1", "Pressure sensor with longer description"],
        ]
        name_col = _find_name_column_by_content(rows, header_row_idx=0, num_cols=3)
        
        # Column 2 has longest average text length
        assert name_col == 2
