"""Загрузчик и парсер таблиц параметров.

Поддерживает:
- Excel .xlsx (основной, UTF-8) — адаптивное определение колонок
- CSV с разделителем ; (fallback, Windows-1251)

Семантическое определение ролей колонок по ключевым словам в заголовках.
Не использует захардкоженные индексы колонок.
"""

from __future__ import annotations

import csv
import io
import logging
import time
from pathlib import Path

import chardet

logger = logging.getLogger(__name__)

# Ключевые слова для определения роли колонки (case-insensitive)
COLUMN_KEYWORDS = {
    'name': ['наименование', 'описание', 'название', 'параметр', 'name', 'имя', 'русское'],
    'unit': ['единиц', 'измерен', 'размерность', 'ед.', 'ед', 'unit', 'dimension'],
    'short_name': ['коротк', 'сокращ', 'short', 'код', 'тег', 'tag', 'обознач', 'тип', 'величина'],
    'decimal_places': ['знак', 'точн', 'разряд', 'decimal', 'знаки', 'разряды'],
    'type': ['тип', 'величина', 'type', 'физ.'],
    'sensor_type': ['датчик', 'sensor'],
    'id': ['№', 'номер', 'id', 'n', 'number'],
}


def load_parameter_table(filepath: str | Path) -> list[dict]:
    """Загружает таблицу параметров из файла.

    Args:
        filepath: Путь к файлу (.xlsx или .csv).

    Returns:
        Список словарей с полями: id, name, unit, short_name, decimal_places, sheet_name.
        Поле sheet_name содержит имя листа Excel (для .xlsx) или пустую строку (для .csv).
    """
    t0 = time.perf_counter()
    filepath = Path(filepath)

    if filepath.suffix.lower() == ".xlsx":
        result = _load_xlsx(filepath)
    elif filepath.suffix.lower() == ".csv":
        result = _load_csv(filepath)
    else:
        logger.error("Неподдерживаемый формат файла: %s", filepath.suffix)
        raise ValueError(f"Неподдерживаемый формат: {filepath.suffix}")

    elapsed_ms = (time.perf_counter() - t0) * 1000
    logger.info("Загружена таблица параметров: %d записей из %s (%.1f мс)", len(result), filepath.name, elapsed_ms)

    return result


def _load_xlsx(filepath: Path) -> list[dict]:
    """Загружает таблицу из Excel .xlsx.

    Обходит ВСЕ листы в книге и собирает параметры в единый список.
    Каждому параметру присваивается уникальный глобальный ID и сохраняется имя листа.

    Использует адаптивное определение колонок по семантическим ролям.

    Args:
        filepath: Путь к .xlsx файлу.

    Returns:
        Список параметров с полями: id, name, unit, short_name, decimal_places, sheet_name.
    """
    from openpyxl import load_workbook

    wb = load_workbook(filepath, read_only=True)
    sheet_names = wb.sheetnames
    logger.debug("Найдено %d листов в файле %s: %s", len(sheet_names), filepath.name, sheet_names)

    params: list[dict] = []
    global_param_id = 0  # Глобальный счётчик ID across all sheets

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        sheet_params_count = 0

        # Получаем все строки листа
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            logger.debug("Лист '%s' пуст, пропускаем", sheet_name)
            continue

        # Определяем строку с заголовками
        header_row_idx = _detect_header_row(rows, max_rows=5)
        logger.debug("Лист '%s': заголовок найден в строке %d", sheet_name, header_row_idx + 1)

        # Получаем заголовки
        headers = [str(h).strip().lower() if h else "" for h in rows[header_row_idx]]

        # Определяем роли колонок по ключевым словам
        column_roles = _detect_column_roles(headers)
        _log_column_roles(sheet_name, column_roles)

        # Если колонка name не найдена — используем fallback
        if column_roles.get('name') is None:
            column_roles['name'] = _find_name_column_by_content(rows, header_row_idx, len(headers))
            if column_roles['name'] is not None:
                logger.debug("Лист '%s': name колонка определена по содержимому: колонка %d",
                             sheet_name, column_roles['name'] + 1)

        # Индексы колонок
        name_col = column_roles.get('name')
        unit_col = column_roles.get('unit')
        short_col = column_roles.get('short_name')
        decimal_col = column_roles.get('decimal_places')
        id_col = column_roles.get('id')

        # Если short_name не найдена, пробуем использовать 'type'
        if short_col is None:
            short_col = column_roles.get('type')

        # Читаем данные начиная со строки после заголовка
        for row in rows[header_row_idx + 1:]:
            if not row or all(v is None for v in row):
                continue

            # Используем порядковый номер как глобальный ID
            global_param_id += 1
            param_id = global_param_id

            # Если в таблице есть свой ID — используем его (но это опционально)
            if id_col is not None and id_col < len(row) and row[id_col] is not None:
                try:
                    _sheet_local_id = int(row[id_col])  # noqa: F841
                except (ValueError, TypeError):
                    pass

            name = str(row[name_col] or "") if name_col is not None and name_col < len(row) else ""
            unit = str(row[unit_col] or "") if unit_col is not None and unit_col < len(row) else ""
            short = str(row[short_col] or "") if short_col is not None and short_col < len(row) else ""
            decimal = 1  # default
            if decimal_col is not None and decimal_col < len(row) and row[decimal_col] is not None:
                try:
                    decimal = int(row[decimal_col])
                except (ValueError, TypeError):
                    decimal = 1

            if not name.strip():
                continue

            params.append({
                "id": param_id,
                "name": name.strip(),
                "unit": unit.strip(),
                "short_name": short.strip(),
                "decimal_places": decimal,
                "sheet_name": sheet_name,
            })
            sheet_params_count += 1

        logger.debug("Лист '%s': загружено %d параметров", sheet_name, sheet_params_count)

    wb.close()
    logger.info(
        "Загружено %d параметров из %d листов (файл: %s)",
        len(params),
        len(sheet_names),
        filepath.name,
    )
    return params


def _detect_header_row(rows: list, max_rows: int = 5) -> int:
    """Определяет строку с заголовками по наличию текстовых ячеек.

    Ищет строку, содержащую минимум 2 текстовых значения длиной > 1 символа.
    Такие строки скорее всего являются заголовками.

    Args:
        rows: Список строк из Excel листа.
        max_rows: Максимальное количество строк для проверки.

    Returns:
        Индекс строки с заголовками (0-based). Если не найдено, возвращает 0.
    """
    for row_idx in range(min(max_rows, len(rows))):
        row = rows[row_idx]
        if not row:
            continue
        # Считаем текстовые ячейки с длиной > 1
        text_cells = sum(
            1 for cell in row
            if cell is not None and isinstance(cell, str) and len(str(cell).strip()) > 1
        )
        # Минимум 2 текстовых колонки = похоже на заголовок
        if text_cells >= 2:
            return row_idx
    return 0  # По умолчанию первая строка


def _detect_column_roles(headers: list[str]) -> dict[str, int | None]:
    """Определяет роли колонок по ключевым словам в заголовках.

    Для каждого заголовка проверяет, содержит ли он ключевые слова
    для какой-либо роли. Возвращает словарь с индексами колонок для каждой роли.

    Args:
        headers: Список заголовков колонок (приведённых к нижнему регистру).

    Returns:
        Словарь {role: column_index} для найденных ролей.
    """
    roles: dict[str, int | None] = {}
    used_columns: set[int] = set()

    # Приоритет ролей при конфликте (более специфичные сначала)
    role_priority = ['id', 'decimal_places', 'short_name', 'sensor_type', 'type', 'unit', 'name']

    for role in role_priority:
        keywords = COLUMN_KEYWORDS.get(role, [])
        for col_idx, header in enumerate(headers):
            if col_idx in used_columns:
                continue
            header_lower = header.lower()
            # Проверяем, содержит ли заголовок ключевое слово
            for keyword in keywords:
                if keyword in header_lower:
                    roles[role] = col_idx
                    used_columns.add(col_idx)
                    break
            if roles.get(role) is not None:
                break

    return roles


def _find_name_column_by_content(rows: list, header_row_idx: int, num_cols: int) -> int | None:
    """Находит колонку name по содержимому (колонка с самым длинным средним текстом).

    Fallback-метод, когда колонка name не найдена по заголовку.

    Args:
        rows: Все строки листа.
        header_row_idx: Индекс строки с заголовками.
        num_cols: Количество колонок.

    Returns:
        Индекс колонки с самым длинным средним текстом или None.
    """
    if num_cols == 0:
        return None

    col_lengths: list[float] = [0.0] * num_cols
    col_counts: list[int] = [0] * num_cols

    # Анализируем до 10 строк данных после заголовка
    sample_rows = rows[header_row_idx + 1:header_row_idx + 11]

    for row in sample_rows:
        for col_idx in range(num_cols):
            if col_idx < len(row) and row[col_idx] is not None:
                cell_value = str(row[col_idx])
                col_lengths[col_idx] += len(cell_value)
                col_counts[col_idx] += 1

    # Находим колонку с максимальной средней длиной
    max_avg = 0.0
    max_col = None
    for col_idx in range(num_cols):
        if col_counts[col_idx] > 0:
            avg_length = col_lengths[col_idx] / col_counts[col_idx]
            if avg_length > max_avg:
                max_avg = avg_length
                max_col = col_idx

    return max_col


def _log_column_roles(sheet_name: str, column_roles: dict[str, int | None]) -> None:
    """Логирует определённые роли колонок.

    Args:
        sheet_name: Имя листа.
        column_roles: Словарь с ролями колонок.
    """
    # Преобразуем индексы в буквенные обозначения колонок (A, B, C, ...)
    def col_letter(idx: int | None) -> str:
        if idx is None:
            return "-"
        result = ""
        idx += 1  # 1-based
        while idx > 0:
            idx -= 1
            result = chr(ord('A') + idx % 26) + result
            idx //= 26
        return result

    parts = []
    for role in ['name', 'unit', 'short_name', 'decimal_places', 'type', 'sensor_type', 'id']:
        if role in column_roles:
            parts.append(f"{role}={col_letter(column_roles[role])}")

    logger.info("Лист '%s': определены колонки — %s", sheet_name, ", ".join(parts))


def _load_csv(filepath: Path) -> list[dict]:
    """Загружает таблицу из CSV (разделитель ;, кодировка auto-detect).

    Использует адаптивное определение колонок по семантическим ролям.

    Args:
        filepath: Путь к CSV файлу.

    Returns:
        Список параметров.
    """
    raw = filepath.read_bytes()

    # Определяем кодировку
    detected = chardet.detect(raw)
    encoding = detected.get("encoding", "windows-1251") or "windows-1251"
    logger.debug("CSV кодировка определена: %s (confidence: %.2f)", encoding, detected.get("confidence", 0))

    text = raw.decode(encoding, errors="replace")
    reader = csv.reader(io.StringIO(text), delimiter=";")

    params: list[dict] = []
    rows = list(reader)

    if not rows:
        logger.warning("CSV файл пуст: %s", filepath.name)
        return params

    # Определяем строку с заголовками
    header_row_idx = _detect_header_row_csv(rows, max_rows=5)
    logger.debug("CSV: заголовок найден в строке %d", header_row_idx + 1)

    # Получаем заголовки
    headers = [str(h).strip().lower() if h else "" for h in rows[header_row_idx]]

    # Определяем роли колонок по ключевым словам
    column_roles = _detect_column_roles(headers)
    _log_column_roles("CSV", column_roles)

    # Если колонка name не найдена — используем fallback
    if column_roles.get('name') is None:
        column_roles['name'] = _find_name_column_by_content_csv(rows, header_row_idx, len(headers))
        if column_roles['name'] is not None:
            logger.debug("CSV: name колонка определена по содержимому: колонка %d",
                         column_roles['name'] + 1)

    # Индексы колонок
    name_col = column_roles.get('name')
    unit_col = column_roles.get('unit')
    id_col = column_roles.get('id')

    row_idx = 0
    for row in rows[header_row_idx + 1:]:
        row_idx += 1
        if not row:
            continue

        param_id = row_idx
        if id_col is not None and id_col < len(row):
            try:
                param_id = int(row[id_col])
            except (ValueError, TypeError):
                # Не числовой ID (например, "PT001") — используем порядковый номер
                pass

        name = row[name_col].strip() if name_col is not None and name_col < len(row) else ""
        unit = row[unit_col].strip() if unit_col is not None and unit_col < len(row) else ""

        if not name:
            continue

        params.append({
            "id": param_id,
            "name": name,
            "unit": unit,
            "short_name": "",
            "decimal_places": 1,
            "sheet_name": "",  # CSV не имеет листов
        })

    logger.debug("Загружено %d параметров из CSV", len(params))
    return params


def _detect_header_row_csv(rows: list, max_rows: int = 5) -> int:
    """Определяет строку с заголовками в CSV.

    Args:
        rows: Список строк из CSV.
        max_rows: Максимальное количество строк для проверки.

    Returns:
        Индекс строки с заголовками (0-based). Если не найдено, возвращает 0.
    """
    for row_idx in range(min(max_rows, len(rows))):
        row = rows[row_idx]
        if not row:
            continue
        # Считаем текстовые ячейки с длиной > 1
        text_cells = sum(
            1 for cell in row
            if cell and len(str(cell).strip()) > 1
        )
        # Минимум 2 текстовых колонки = похоже на заголовок
        if text_cells >= 2:
            return row_idx
    return 0  # По умолчанию первая строка


def _find_name_column_by_content_csv(rows: list, header_row_idx: int, num_cols: int) -> int | None:
    """Находит колонку name в CSV по содержимому.

    Fallback-метод, когда колонка name не найдена по заголовку.

    Args:
        rows: Все строки CSV.
        header_row_idx: Индекс строки с заголовками.
        num_cols: Количество колонок.

    Returns:
        Индекс колонки с самым длинным средним текстом или None.
    """
    if num_cols == 0:
        return None

    col_lengths: list[float] = [0.0] * num_cols
    col_counts: list[int] = [0] * num_cols

    # Анализируем до 10 строк данных после заголовка
    sample_rows = rows[header_row_idx + 1:header_row_idx + 11]

    for row in sample_rows:
        for col_idx in range(num_cols):
            if col_idx < len(row) and row[col_idx]:
                cell_value = str(row[col_idx])
                col_lengths[col_idx] += len(cell_value)
                col_counts[col_idx] += 1

    # Находим колонку с максимальной средней длиной
    max_avg = 0.0
    max_col = None
    for col_idx in range(num_cols):
        if col_counts[col_idx] > 0:
            avg_length = col_lengths[col_idx] / col_counts[col_idx]
            if avg_length > max_avg:
                max_avg = avg_length
                max_col = col_idx

    return max_col
